import openpyxl as xl
from openpyxl import Workbook
import pandas as pd
from openpyxl.styles import Alignment
import win32com.client as win32
import sys


win32c = win32.constants


def dni(date):
    wb = xl.load_workbook(filename='C:/Users/a4583/OneDrive/Desktop/work/продажи.xlsx')
    ws = wb['Sheet']
    wb1 = Workbook()
    ws1 = wb1.active
    pos_list = []
    count_list = []
    ss_list = []
    sell_list = []
    sold_list = []
    for i in range(2, ws.max_row+1):
        pos = ws.cell(row=i, column=1).value
        if pos not in pos_list:
            pos_list.append(pos)
            count_list.append(0)
            ss_list.append(0)
            sell_list.append(0)
            sold_list.append(0)
    b = 3
    for i in range(3, len(pos_list)):
        z = 'B' + str(b)
        x = 'C' + str(b)
        ws1[z] = int('0')
        ws1[x] = int('0')
        b += 1
    b = 3
    for i in range(2, ws.max_row+1):
        pos = ws.cell(row=i, column=1).value
        pos_index = pos_list.index(pos)
        try:
            count_list[pos_index] += ws.cell(row=i, column=4).value
            ss_list[pos_index] += ws.cell(row=i, column=8).value
        except TypeError:
            print()
    for i in range(1, ws.max_row+1):
        pos = ws.cell(row=i, column=1).value
        try:
            pos = pos.capitalize()
        except AttributeError:
            print()
        if pos in pos_list:
            pos_index = pos_list.index(pos)
            try:
                sell_list[pos_index] += ws.cell(row=i, column=10).value
                sold_list[pos_index] += ws.cell(row=i, column=9).value
            except TypeError:
                print()
    for i in range(3, len(pos_list)):
        count_cell = 'B' + str(i)
        ss_cell = 'C' + str(i)
        pos_cell = 'A' + str(i)
        sell_cell = 'E' + str(i)
        sold_cell = 'D' + str(i)
        ws1[pos_cell] = pos_list[i-3]
        ws1[count_cell] = count_list[i-3]
        ws1[ss_cell] = ss_list[i-3]
        ws1[sold_cell] = float(sold_list[i-3]) / float(count_list[i-3])
        try:
            ws1[sell_cell] = float(sell_list[i - 3]) / float(ss_list[i - 3])
        except ZeroDivisionError:
            print()
        b += 1
        ws1['B1'] = 'Пришло'
    wb1.save('C:/Users/a4583/OneDrive/Desktop/work/продажи_по_дням_сбор ' + date + '.xlsx')


def sort(date):
    wb = pd.read_excel('C:/Users/a4583/OneDrive/Desktop/work/продажи_по_дням_сбор ' + date + '.xlsx')
    wb = wb.sort_values(by='Пришло', ascending=False)
    wb.to_excel('C:/Users/a4583/OneDrive/Desktop/work/продажи_по_дням_сбор ' + date + '.xlsx', index=False)
    wb = xl.load_workbook(filename='C:/Users/a4583/OneDrive/Desktop/work/продажи_по_дням_сбор ' + date + '.xlsx')
    ws = wb.active
    ws['A1'] = ''
    ws['B1'] = ''
    ws['C1'] = ''
    ws['D1'] = ''
    ws['E1'] = ''
    ws.move_range('A1:E500', rows=1, cols=0)
    wb.save('C:/Users/a4583/OneDrive/Desktop/work/продажи_по_дням_сбор ' + date + '.xlsx')


def copy(date, date1, date2):
    wb = xl.load_workbook('C:/Users/a4583/OneDrive/Desktop/work/продажи_по_дням_сбор ' + date + '.xlsx')
    wb2 = xl.load_workbook('C:/Users/a4583/OneDrive/Desktop/work/продажи_по_дням_сбор ' + date + '.xlsx')
    ws = wb.active
    # noinspection PyBroadException
    try:
        wb1 = xl.load_workbook('C:/Users/a4583/OneDrive/Desktop/work/по дням с ' + date1 + '.xlsx')
    except Exception:
        wb1 = wb2
        ws1 = wb1.active
        ws1.delete_cols(4, 5)
    ws1 = wb1.active
    col = ws1.max_column + 1
    ws1.cell(row=1, column=col).value = date2
    for i in range(3, ws.max_row + 1):
        ws1.cell(row=i, column=col).value = ws.cell(row=i, column=4).value
        ws1.cell(row=i, column=col).number_format = '0.00%'
        ws1.cell(row=i, column=col + 1).value = ws.cell(row=i, column=5).value
        ws1.cell(row=i, column=col + 1).number_format = '0.00%'
    ws1.cell(row=2, column=col).value = 'Продано %'
    ws1.cell(row=2, column=col + 1).value = 'Окупаемость %'
    wb1.save('C:/Users/a4583/OneDrive/Desktop/work/по дням с ' + date1 + '.xlsx')


def run_excel(filename: str):
    wb = xl.load_workbook(filename)
    ws = wb.active
    col = ws.max_column
    d = {5: 'E',
         7: 'G',
         9: 'I',
         11: 'K',
         13: 'M',
         15: 'O',
         17: 'Q',
         19: 'S',
         21: 'U',
         23: 'W',
         25: 'Y',
         27: 'AA',
         29: 'AC',
         31: 'AE',
         33: 'AG',
         35: 'AI',
         37: 'AK',
         39: 'AM',
         41: 'AO',
         43: 'AQ',
         45: 'AS',
         47: 'AU',
         49: 'AW',
         51: 'AY',
         53: 'BA',
         55: 'BC',
         57: 'BE',
         59: 'BG'}
    d_col = d[col]

    # create excel object
    excel = win32.gencache.EnsureDispatch('Excel.Application')

    # excel can be visible or not
    excel.Visible = False

    # try except for file / path
    try:
        wb = excel.Workbooks.Open(filename)
    except pywintypes.com_error as e:
        if e.excepinfo[5] == -2146827284:
            print(f'Failed to open spreadsheet.  Invalid filename or location: {filename}')
        else:
            raise e
        sys.exit(1)

    # set worksheet
    ws1 = wb.Sheets('Sheet1')

    ws1.Range(d_col + ":" + d_col).Select()
    excel.Selection.FormatConditions.Add(win32c.xlCellValue, win32c.xlGreater, "=1")
    excel.Selection.FormatConditions(excel.Selection.FormatConditions.Count).SetFirstPriority()
    csc2 = excel.Selection.FormatConditions(1).Font
    csc2.Color = -16752384
    csc2.TintAndShade = 0
    csc3 = excel.Selection.FormatConditions(1).Interior
    csc3.PatternColorIndex = win32c.xlAutomatic
    csc3.Color = 13561798
    csc3.TintAndShade = 0
    excel.Selection.FormatConditions(1).StopIfTrue = False
    excel.Selection.FormatConditions.Add(win32c.xlCellValue, win32c.xlLess, "=1")
    excel.Selection.FormatConditions(excel.Selection.FormatConditions.Count).SetFirstPriority()
    csc5 = excel.Selection.FormatConditions(1).Font
    csc5.Color = -16383844
    csc5.TintAndShade = 0
    csc6 = excel.Selection.FormatConditions(1).Interior
    csc6.PatternColorIndex = win32c.xlAutomatic
    csc6.Color = 13551615
    csc6.TintAndShade = 0
    excel.Selection.FormatConditions(1).StopIfTrue = False
    wb.Close(True)
    excel.Quit()
    wb = xl.load_workbook(filename)
    ws = wb.active
    col = ws.max_column
    ws.merge_cells(start_row=1, start_column=col - 1, end_row=1, end_column=col)
    ws.cell(row=1, column=col - 1).alignment = Alignment(horizontal='center')
    wb.save(filename)
