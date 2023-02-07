import openpyxl as xl


def graph(date1):
    wb = xl.load_workbook(filename='C:/Users/a4583/OneDrive/Desktop/work/графики.xlsx')
    ws = wb['Лист1']
    wb1 = xl.load_workbook('C:/Users/a4583/OneDrive/Desktop/work/по дням с ' + date1 + '.xlsx')
    ws1 = wb1.active
    pos = 2

    def copy(naim):
        for i in range(3, ws1.max_row + 1):
            a1 = ws.cell(row=2, column=naim).value
            a2 = ws1.cell(row=i, column=1).value
            if a1 == a2:
                for a in range(3, ws.max_row + 1):
                    b1 = ws.cell(row=a, column=1).value
                    if b1 == date1:
                        for b in range(naim + 2, naim + 37):
                            c1 = type(ws.cell(row=a, column=b).value)
                            c2 = type(ws.cell(row=1, column=1).value)
                            if c1 == c2:
                                ws.cell(row=a, column=b).value = ws1.cell(row=i, column=ws1.max_column).value
                                ws.cell(row=a, column=b).number_format = '0.00%'
                                break

    while ws.cell(row=2, column=pos).value is not None:
        copy(pos)
        pos += 38
    wb.save(filename='C:/Users/a4583/OneDrive/Desktop/work/графики.xlsx')
