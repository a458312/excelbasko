import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import win32com.client as win32
import sys

win32c = win32.constants


def prod(date):
    def openworkbook(xlapp, xlfile):
        try:
            xlwb = xlapp.Workbooks(xlfile)
        except Exception as e:
            try:
                xlwb = xlapp.Workbooks.Open(xlfile)
            except Exception as e:
                print(e)
                xlwb = None
        return (xlwb)

    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = openworkbook(excel, 'C:/Users/a4583/OneDrive/Desktop/work/prodazhitest' + date + '.xlsx')
        ws = wb.Worksheets('TDSheet')
        excel.Visible = False
        excel.ActiveWorkbook.Save()

    except Exception as e:
        print(e)

    finally:
        # RELEASES RESOURCES
        ws = None
        wb = None
        excel = None

    wb = xl.load_workbook(filename='C:/Users/a4583/OneDrive/Desktop/work/prihodtest' + date + '.xlsx')
    wb2 = xl.load_workbook(filename='C:/Users/a4583/OneDrive/Desktop/work/prodazhitest' + date + '.xlsx')
    ws = wb.active
    ws2 = wb2.active
    wb1 = Workbook()
    ws1 = wb1.active
    pos_list = [0, 0, 0, 0, 0, 0, 0, 0]
    sell_list = [0, 0, 0, 0, 0]
    b = 2
    for i in range(13, ws.max_row + 1):
        font = ws.cell(row=i, column=2).font
        try:
            font2 = font.color.rgb
        except AttributeError:
            font = Font(color='FF000001')
        if font.bold and font.color.rgb == 'FF594304':
            pos_list[0] = ws.cell(row=i, column=2).value  # mag
            pos_list[0] = pos_list[0].replace(' подготовительный', '')
            pos_list[0] = pos_list[0].replace('ЧЛБ ', '')
            pos_list[0] = pos_list[0].replace('К-УР ', '')
            pos_list[0] = pos_list[0].replace('Пермь ', '')
        if font.bold and font.color.rgb == 'FF000000':
            pos_list[1] = ws.cell(row=i, column=2).value  # art
            try:
                pos_list[6] = ws.cell(row=i, column=5).value/ws.cell(row=i, column=4).value  # sskg
            except TypeError:
                print()
        if not font.bold and font.color.rgb == 'FF000000':
            pos_list[2] = ws.cell(row=i, column=2).value  # naim
            pos_list[2] = pos_list[2].capitalize()
            pos_list[3] = ws.cell(row=i, column=3).value  # kolvo
            pos_list[4] = ws.cell(row=i, column=4).value  # ves rsch
            pos_list[5] = pos_list[4] / pos_list[3]  # ves sred
            pos_list[7] = ws.cell(row=i, column=5).value  # ss rsch
            pos_cell = 'A' + str(b)
            art_cell = 'B' + str(b)
            shop_cell = 'C' + str(b)
            count_cell = 'D' + str(b)
            weight_cell = 'E' + str(b)
            sold_percent_cell = 'K' + str(b)
            marja_cell = 'L' + str(b)
            total_weight_cell = 'F' + str(b)
            ss_cell = 'G' + str(b)
            ss_calc_cell = 'H' + str(b)
            ws1[pos_cell] = pos_list[2]
            ws1[art_cell] = pos_list[1]
            ws1[shop_cell] = pos_list[0]
            ws1[count_cell] = pos_list[3]
            ws1[weight_cell] = pos_list[5]
            ws1[total_weight_cell] = pos_list[4]
            ws1[ss_cell] = pos_list[6]
            ws1[ss_calc_cell] = pos_list[7]
            ws1[sold_percent_cell] = '=I' + str(b) + '/D' + str(b)
            ws1[marja_cell] = '=J' + str(b) + '/H' + str(b)
            b += 1
    for y in range(11, ws2.max_row + 1):
        font = ws2.cell(row=y, column=2).font
        try:
            font2 = font.color.rgb
        except AttributeError:
            font = Font(color='FF000001')
        if font.bold and font.color.rgb == 'FF594304':
            sell = ws2.cell(row=y, column=2).value  # mag
            sell = sell.replace('Магазин ', '')
            sell = sell.replace('Каменск-Уральский', '')
            sell = sell.replace('Челябинск', '')
            sell = sell.replace('Пермь', '')
            sell = sell.replace(', ', '')
            sell_list[0] = sell
        if font.bold and font.color.rgb == 'FF000000':
            sell_list[1] = ws2.cell(row=y, column=2).value  # art
        if not font.bold and font.color.rgb == 'FF000000':
            sell_list[2] = ws2.cell(row=y, column=2).value  # naim
            sell_list[2] = sell_list[2].capitalize()
            sell_list[3] = ws2.cell(row=y, column=3).value  # summa
            sell_list[4] = ws2.cell(row=y, column=4).value  # kolvo prod
            for z in range(2, ws2.max_row + 1):
                if sell_list[0] == ws1.cell(row=z, column=3).value and \
                        sell_list[1] == ws1.cell(row=z, column=2).value and \
                        sell_list[2] == ws1.cell(row=z, column=1).value:
                    sell_count_cell = 'I' + str(z)
                    price_cell = 'J' + str(z)
                    ws1[sell_count_cell] = sell_list[3]
                    ws1[price_cell] = sell_list[4]
                    break
    ws1['A1'] = 'Наименование'
    ws1['A1'].font = Font(bold=True, size=12)
    ws1['B1'] = 'Артикул'
    ws1['B1'].font = Font(bold=True, size=12)
    ws1['D1'] = 'Пришло'
    ws1['D1'].font = Font(bold=True, size=12)
    ws1['C1'] = 'Магазин'
    ws1['C1'].font = Font(bold=True, size=12)
    ws1['E1'] = 'Средний вес'
    ws1['E1'].font = Font(bold=True, size=12)
    ws1['F1'] = 'Расчетный вес партии'
    ws1['F1'].font = Font(bold=True, size=12)
    ws1['G1'] = 'С/С/кг'
    ws1['G1'].font = Font(bold=True, size=12)
    ws1['H1'] = 'С/С Расчетное'
    ws1['H1'].font = Font(bold=True, size=12)
    ws1['I1'] = 'Продано'
    ws1['I1'].font = Font(bold=True, size=12)
    ws1['J1'] = 'Выручка'
    ws1['J1'].font = Font(bold=True, size=12)
    ws1['K1'] = 'Продано %'
    ws1['K1'].font = Font(bold=True, size=12)
    ws1['L1'] = 'Окупаемость %'
    ws1['L1'].font = Font(bold=True, size=12)
    wb1.save('C:/Users/a4583/OneDrive/Desktop/work/продажитест.xlsx')


prod('1111')
