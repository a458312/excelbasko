import openpyxl as xl
from openpyxl.styles import Font
from openpyxl import Workbook
import win32com.client as win32
import sys


win32c = win32.constants


def prod(date):
    def openworkbook(xlapp, xlfile):
        try:
            xlwb = xlapp.Workbooks(xlfile)
        except Exception as e:
            try:
                xlwb = xlapp.Workbooks.Open(xlfile)
            except Exception as e:
                print(e)
                xlwb = None
        return xlwb

    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = openworkbook(excel, 'C:/Users/a4583/OneDrive/Desktop/work/prodazhi' + date + '.xlsx')
        ws = wb.Worksheets('TDSheet')
        excel.Visible = False
        excel.ActiveWorkbook.Save()

    except Exception as e:
        print(e)

    finally:
        # RELEASES RESOURCES
        ws = None
        wb = None
        excel = None

    wb = xl.load_workbook(filename='C:/Users/a4583/OneDrive/Desktop/work/prihod' + date + '.xlsx')
    wb2 = xl.load_workbook(filename='C:/Users/a4583/OneDrive/Desktop/work/prodazhi' + date + '.xlsx')
    ws = wb.active
    ws2 = wb2.active
    wb1 = Workbook()
    ws1 = wb1.active
    pos_list = [0, 0, 0, 0, 0, 0, 0, 0]
    sell_list = [0, 0, 0, 0, 0]
    b = 2
    for i in range(13, ws.max_row + 1):
        font = ws.cell(row=i, column=2).font
        try:
            font2 = font.color.rgb
        except AttributeError:
            font = Font(color='FF000001')
        if font.bold and font.color.rgb == 'FF594304':
            pos_list[0] = ws.cell(row=i, column=2).value  # mag
            pos_list[0] = pos_list[0].replace(' подготовительный', '')
            pos_list[0] = pos_list[0].replace('ЧЛБ ', '')
            pos_list[0] = pos_list[0].replace('К-УР ', '')
            pos_list[0] = pos_list[0].replace('Пермь ', '')
        if font.bold and font.color.rgb == 'FF000000':
            pos_list[1] = ws.cell(row=i, column=2).value  # art
            try:
                pos_list[6] = ws.cell(row=i, column=5).value / ws.cell(row=i, column=4).value  # sskg
            except TypeError:
                print()
        if not font.bold and font.color.rgb == 'FF000000':
            pos_list[2] = ws.cell(row=i, column=2).value  # naim
            pos_list[2] = pos_list[2].capitalize()
            pos_list[3] = ws.cell(row=i, column=3).value  # kolvo
            pos_list[4] = ws.cell(row=i, column=4).value  # ves rsch
            try:
                pos_list[5] = pos_list[4] / pos_list[3]  # ves sred
            except TypeError:
                print()
            pos_list[7] = ws.cell(row=i, column=5).value  # ss rsch
            pos_cell = 'A' + str(b)
            art_cell = 'B' + str(b)
            shop_cell = 'C' + str(b)
            count_cell = 'D' + str(b)
            weight_cell = 'E' + str(b)
            sold_percent_cell = 'K' + str(b)
            marja_cell = 'L' + str(b)
            total_weight_cell = 'F' + str(b)
            ss_cell = 'G' + str(b)
            ss_calc_cell = 'H' + str(b)
            ws1[pos_cell] = pos_list[2]
            ws1[art_cell] = pos_list[1]
            ws1[shop_cell] = pos_list[0]
            ws1[count_cell] = pos_list[3]
            ws1[weight_cell] = pos_list[5]
            ws1[total_weight_cell] = pos_list[4]
            ws1[ss_cell] = pos_list[6]
            ws1[ss_calc_cell] = pos_list[7]
            ws1[sold_percent_cell] = '=I' + str(b) + '/D' + str(b)
            ws1[marja_cell] = '=J' + str(b) + '/H' + str(b)
            b += 1
    for y in range(11, ws2.max_row + 1):
        font = ws2.cell(row=y, column=2).font
        try:
            font2 = font.color.rgb
        except AttributeError:
            font = Font(color='FF000001')
        if font.bold and font.color.rgb == 'FF594304':
            sell = ws2.cell(row=y, column=2).value  # mag
            sell = sell.replace('Магазин ', '')
            sell = sell.replace('Каменск-Уральский', '')
            sell = sell.replace('Челябинск', '')
            sell = sell.replace('Пермь', '')
            sell = sell.replace(', ', '')
            sell_list[0] = sell
        if font.bold and font.color.rgb == 'FF000000':
            sell_list[1] = ws2.cell(row=y, column=2).value  # art
        if not font.bold and font.color.rgb == 'FF000000':
            sell_list[2] = ws2.cell(row=y, column=2).value  # naim
            sell_list[2] = sell_list[2].capitalize()
            sell_list[3] = ws2.cell(row=y, column=3).value  # summa
            sell_list[4] = ws2.cell(row=y, column=4).value  # kolvo prod
            for z in range(2, ws1.max_row + 1):
                if sell_list[0] == ws1.cell(row=z, column=3).value and \
                        sell_list[1] == ws1.cell(row=z, column=2).value and \
                        sell_list[2] == ws1.cell(row=z, column=1).value:
                    sell_count_cell = 'I' + str(z)
                    price_cell = 'J' + str(z)
                    ws1[sell_count_cell] = sell_list[3]
                    ws1[price_cell] = sell_list[4]
                    break
    ws1['A1'] = 'Наименование'
    ws1['A1'].font = Font(bold=True, size=12)
    ws1['B1'] = 'Артикул'
    ws1['B1'].font = Font(bold=True, size=12)
    ws1['D1'] = 'Пришло'
    ws1['D1'].font = Font(bold=True, size=12)
    ws1['C1'] = 'Магазин'
    ws1['C1'].font = Font(bold=True, size=12)
    ws1['E1'] = 'Средний вес'
    ws1['E1'].font = Font(bold=True, size=12)
    ws1['F1'] = 'Расчетный вес партии'
    ws1['F1'].font = Font(bold=True, size=12)
    ws1['G1'] = 'С/С/кг'
    ws1['G1'].font = Font(bold=True, size=12)
    ws1['H1'] = 'С/С Расчетное'
    ws1['H1'].font = Font(bold=True, size=12)
    ws1['I1'] = 'Продано'
    ws1['I1'].font = Font(bold=True, size=12)
    ws1['J1'] = 'Выручка'
    ws1['J1'].font = Font(bold=True, size=12)
    ws1['K1'] = 'Продано %'
    ws1['K1'].font = Font(bold=True, size=12)
    ws1['L1'] = 'Окупаемость %'
    ws1['L1'].font = Font(bold=True, size=12)
    wb1.save('C:/Users/a4583/OneDrive/Desktop/work/продажи.xlsx')


def run_excel(filename: str):

    # create excel object
    excel = win32.gencache.EnsureDispatch('Excel.Application')

    # excel can be visible or not
    excel.Visible = False

    # try except for file / path
    try:
        wb = excel.Workbooks.Open(filename)
    except pywintypes.com_error as e:
        if e.excepinfo[5] == -2146827284:
            print(f'Failed to open spreadsheet.  Invalid filename or location: {filename}')
        else:
            raise e
        sys.exit(1)

    # set worksheet
    ws1 = wb.Sheets('Sheet')

    # Setup and call pivot_table
    ws2_name = 'pivot_table'
    wb.Sheets.Add().Name = ws2_name
    ws2 = wb.Sheets(ws2_name)

    pt_name = 'example'  # must be a string
    pt_rows = ['Магазин', 'Артикул', 'Наименование']  # must be a list
    pt_cols = []  # must be a list
    pt_filters = []  # must be a list
    # [0]: field name [1]: pivot table column name [3]: calculation method [4]: number format
    pt_fields = [['Пришло', 'Sum of Пришло', win32c.xlSum, '0,00'],  # must be a list of lists
                 ['С/С Расчетное', 'Sum of С/С Расчетное', win32c.xlSum, '0,00'],
                 ['Продано', 'Sum of Продано', win32c.xlSum, '0,00'],
                 ['Выручка', 'Sum of Выручка', win32c.xlSum, '0,00']]

    pivot_table(wb, ws1, ws2, ws2_name, pt_name, pt_rows, pt_cols, pt_filters, pt_fields)
    ws2.Range("G:G").Select()
    excel.Selection.FormatConditions.Add(win32c.xlCellValue, win32c.xlGreater, "=1")
    excel.Selection.FormatConditions(excel.Selection.FormatConditions.Count).SetFirstPriority()
    csc2 = excel.Selection.FormatConditions(1).Font
    csc2.Color = -16752384
    csc2.TintAndShade = 0
    csc3 = excel.Selection.FormatConditions(1).Interior
    csc3.PatternColorIndex = win32c.xlAutomatic
    csc3.Color = 13561798
    csc3.TintAndShade = 0
    excel.Selection.FormatConditions(1).StopIfTrue = False
    excel.Selection.FormatConditions.Add(win32c.xlCellValue, win32c.xlLess, "=1")
    excel.Selection.FormatConditions(excel.Selection.FormatConditions.Count).SetFirstPriority()
    csc5 = excel.Selection.FormatConditions(1).Font
    csc5.Color = -16383844
    csc5.TintAndShade = 0
    csc6 = excel.Selection.FormatConditions(1).Interior
    csc6.PatternColorIndex = win32c.xlAutomatic
    csc6.Color = 13551615
    csc6.TintAndShade = 0
    excel.Selection.FormatConditions(1).StopIfTrue = False
    wb.Close(True)
    excel.Quit()


def pivot_table(wb: object, ws1: object, pt_ws: object, ws_name: str, pt_name: str, pt_rows: list, pt_cols: list,
                pt_filters: list, pt_fields: list):
    """
    wb = workbook1 reference
    ws1 = worksheet1
    pt_ws = pivot table worksheet number
    ws_name = pivot table worksheet name
    pt_name = name given to pivot table
    pt_rows, pt_cols, pt_filters, pt_fields: values selected for filling the pivot tables
    """

    # pivot table location
    pt_loc = len(pt_filters) + 2

    # grab the pivot table source data
    pc = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=ws1.UsedRange)

    # create the pivot table object
    pc.CreatePivotTable(TableDestination=f'{ws_name}!R{pt_loc}C1', TableName=pt_name)

    # select the pivot table work sheet and location to create the pivot table
    pt_ws.Select()
    pt_ws.Cells(pt_loc, 1).Select()

    # Sets the rows, columns and filters of the pivot table
    for field_list, field_r in (
                                (pt_filters, win32c.xlPageField),
                                (pt_rows, win32c.xlRowField),
                                (pt_cols, win32c.xlColumnField)):
        for i, value in enumerate(field_list):
            pt_ws.PivotTables(pt_name).PivotFields(value).Orientation = field_r
            pt_ws.PivotTables(pt_name).PivotFields(value).Position = i + 1

    # Sets the Values of the pivot table
    for field in pt_fields:
        pt_ws.PivotTables(pt_name).AddDataField(pt_ws.PivotTables(pt_name).PivotFields(field[0]), field[1],
                                                field[2]).NumberFormat = field[3]

    # Visibility True or false
    pt_ws.PivotTables(pt_name).ShowValuesRow = True
    pt_ws.PivotTables(pt_name).ColumnGrand = True


def perc(name):
    wb = xl.load_workbook('C:/Users/a4583/OneDrive/Desktop/work/продажи.xlsx')
    ws = wb['pivot_table']
    ws['F3'] = 'Продано %'
    ws['F3'].font = Font(bold=True, size=12)
    ws['G3'] = 'Окупаемость %'
    ws['G3'].font = Font(bold=True, size=12)
    for i in range(4, ws.max_row+1):
        ws.cell(row=i, column=6).value = '=d'+str(i)+'/b'+str(i)
        ws.cell(row=i, column=6).number_format = '0.00%'
        ws.cell(row=i, column=7).value = '=e'+str(i)+'/c'+str(i)
        ws.cell(row=i, column=7).number_format = '0.00%'
    wb.save('C:/Users/a4583/OneDrive/Desktop/work/продажи.xlsx')
    wb.save('C:/Users/a4583/OneDrive/Desktop/work/продажи ' + name + '.xlsx')
